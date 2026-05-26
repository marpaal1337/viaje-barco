from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from copy import copy

wb = Workbook()

# ── Colores ──
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=10)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
LIGHT_GRAY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

def style_row(ws, row, cols, fill=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if fill:
            cell.fill = fill

def auto_width(ws, cols, min_w=10, max_w=25):
    for c in range(1, cols + 1):
        letter = get_column_letter(c)
        best = min_w
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=False):
            for cell in row:
                if cell.value:
                    best = max(best, min(len(str(cell.value)) + 2, max_w))
        ws.column_dimensions[letter].width = best

# ════════════════════════════════════════════
# HOJA 1: VUELOS
# ════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Vuelos"
ws1.sheet_properties.tabColor = "1F4E79"

# Datos: (Origen, Destino, Aerolínea, Semana, Precio Ida, Precio Vuelta, Total, Notas)
# Semanas: Ago S1 (1-7), Ago S2 (8-14), Ago S3 (15-21), Ago S4 (22-31),
#          Sep S1 (1-7), Sep S2 (8-14), Sep S3 (15-21), Sep S4 (22-30)
flight_data = [
    # VLC → IBZ
    ("Valencia (VLC)", "Ibiza (IBZ)", "Ryanair", "Ago S1", 42, 42, 84, "Vuelo nocturno, sin maleta facturada"),
    ("Valencia (VLC)", "Ibiza (IBZ)", "Ryanair", "Ago S2", 45, 45, 90, ""),
    ("Valencia (VLC)", "Ibiza (IBZ)", "Ryanair", "Ago S3", 48, 48, 96, "Fin de semana más caro"),
    ("Valencia (VLC)", "Ibiza (IBZ)", "Ryanair", "Ago S4", 42, 42, 84, ""),
    ("Valencia (VLC)", "Ibiza (IBZ)", "Ryanair", "Sep S1", 42, 42, 84, "Opción más barata"),
    ("Valencia (VLC)", "Ibiza (IBZ)", "Ryanair", "Sep S2", 38, 38, 76, "Mínimo estimado"),
    ("Valencia (VLC)", "Ibiza (IBZ)", "Ryanair", "Sep S3", 35, 35, 70, "Baja temporada"),
    ("Valencia (VLC)", "Ibiza (IBZ)", "Ryanair", "Sep S4", 35, 35, 70, "Baja temporada"),
    # VLC → MAH
    ("Valencia (VLC)", "Menorca (MAH)", "Ryanair", "Ago S1", 71, 60, 131, ""),
    ("Valencia (VLC)", "Menorca (MAH)", "Ryanair", "Ago S2", 75, 65, 140, ""),
    ("Valencia (VLC)", "Menorca (MAH)", "Ryanair", "Ago S3", 77, 65, 142, ""),
    ("Valencia (VLC)", "Menorca (MAH)", "Ryanair", "Ago S4", 71, 60, 131, ""),
    ("Valencia (VLC)", "Menorca (MAH)", "Ryanair", "Sep S1", 55, 50, 105, "Estimado"),
    ("Valencia (VLC)", "Menorca (MAH)", "Ryanair", "Sep S2", 50, 45, 95, "Estimado"),
    ("Valencia (VLC)", "Menorca (MAH)", "Ryanair", "Sep S3", 45, 45, 90, "Estimado"),
    ("Valencia (VLC)", "Menorca (MAH)", "Ryanair", "Sep S4", 45, 40, 85, "Estimado"),
    # MAD → IBZ
    ("Madrid (MAD)", "Ibiza (IBZ)", "Iberia Express", "Ago S1", 56, 56, 112, "Desde €112 ida+vuelta"),
    ("Madrid (MAD)", "Ibiza (IBZ)", "Ryanair", "Ago S1", 65, 65, 130, ""),
    ("Madrid (MAD)", "Ibiza (IBZ)", "Iberia Express", "Ago S2", 60, 60, 120, "Estimado"),
    ("Madrid (MAD)", "Ibiza (IBZ)", "Ryanair", "Ago S3", 70, 70, 140, "Fin de semana"),
    ("Madrid (MAD)", "Ibiza (IBZ)", "Ryanair", "Sep S1", 35, 32, 67, "Desde €67 ida+vuelta"),
    ("Madrid (MAD)", "Ibiza (IBZ)", "Ryanair", "Sep S2", 35, 35, 70, "Estimado"),
    ("Madrid (MAD)", "Ibiza (IBZ)", "Ryanair", "Sep S3", 30, 30, 60, "Estimado"),
    ("Madrid (MAD)", "Ibiza (IBZ)", "Ryanair", "Sep S4", 30, 30, 60, "Estimado"),
    # MAD → MAH
    ("Madrid (MAD)", "Menorca (MAH)", "Ryanair", "Ago S1", 113, 113, 226, ""),
    ("Madrid (MAD)", "Menorca (MAH)", "Ryanair", "Ago S4", 65, 64, 129, "Desde €129 en fechas mix"),
    ("Madrid (MAD)", "Menorca (MAH)", "Ryanair", "Sep S1", 55, 50, 105, "Estimado"),
    ("Madrid (MAD)", "Menorca (MAH)", "Ryanair", "Sep S2", 50, 45, 95, "Estimado"),
    ("Madrid (MAD)", "Menorca (MAH)", "Ryanair", "Sep S3", 45, 45, 90, "Estimado"),
    ("Madrid (MAD)", "Menorca (MAH)", "Ryanair", "Sep S4", 45, 40, 85, "Estimado"),
    # BLQ → IBZ
    ("Bolonia (BLQ)", "Ibiza (IBZ)", "Ryanair", "Ago S1", 107, 107, 214, "Directo 2h"),
    ("Bolonia (BLQ)", "Ibiza (IBZ)", "Ryanair", "Ago S2", 110, 110, 220, "Estimado"),
    ("Bolonia (BLQ)", "Ibiza (IBZ)", "Ryanair", "Ago S3", 115, 115, 230, "Estimado"),
    ("Bolonia (BLQ)", "Ibiza (IBZ)", "Ryanair", "Ago S4", 107, 107, 214, "Estimado"),
    ("Bolonia (BLQ)", "Ibiza (IBZ)", "Ryanair", "Sep S1", 80, 75, 155, "Estimado"),
    ("Bolonia (BLQ)", "Ibiza (IBZ)", "Ryanair", "Sep S2", 75, 70, 145, "Estimado"),
    ("Bolonia (BLQ)", "Ibiza (IBZ)", "Ryanair", "Sep S3", 70, 65, 135, "Estimado"),
    ("Bolonia (BLQ)", "Ibiza (IBZ)", "Ryanair", "Sep S4", 65, 60, 125, "Estimado"),
    # BLQ → MAH
    ("Bolonia (BLQ)", "Menorca (MAH)", "Vueling", "Ago S1", 143, 143, 286, "1 escala en BCN"),
    ("Bolonia (BLQ)", "Menorca (MAH)", "Vueling", "Sep S1", 100, 95, 195, "Estimado, 1 escala"),
    ("Bolonia (BLQ)", "Menorca (MAH)", "Vueling", "Sep S2", 90, 85, 175, "Estimado, 1 escala"),
    ("Bolonia (BLQ)", "Menorca (MAH)", "Vueling", "Sep S3", 85, 80, 165, "Estimado, 1 escala"),
    ("Bolonia (BLQ)", "Menorca (MAH)", "Vueling", "Sep S4", 80, 75, 155, "Estimado, 1 escala"),
]

headers_v = ["Origen", "Destino", "Aerolínea", "Semana", "Ida (€)", "Vuelta (€)", "Total (€)", "Notas"]
ws1.append(headers_v)
style_header(ws1, 1, len(headers_v))

for i, row in enumerate(flight_data, start=2):
    ws1.append(list(row))
    f = LIGHT_GRAY if i % 2 == 0 else None
    style_row(ws1, i, len(headers_v), fill=f)
    # Color coding for prices
    total = row[6]
    if total <= 80:
        ws1.cell(row=i, column=7).fill = GREEN_FILL
    elif total <= 130:
        ws1.cell(row=i, column=7).fill = YELLOW_FILL
    else:
        ws1.cell(row=i, column=7).fill = RED_FILL

auto_width(ws1, len(headers_v))

# ════════════════════════════════════════════
# HOJA 2: BARCOS
# ════════════════════════════════════════════
ws2 = wb.create_sheet("Barcos")
ws2.sheet_properties.tabColor = "2E75B6"

boat_data = [
    # (Isla, Tipo, Modelo, Eslora, Plazas, Puerto Base, Precio/día Ago, Precio/día Sep, Extras/día, Enlace)
    ("Ibiza", "Velero", "Bavaria 40", "12.5m", 8, "Marina Ibiza", 650, 450, "60 (seguro+combustible)", "clickandboat.com"),
    ("Ibiza", "Velero", "Dufour 390", "11.9m", 8, "Club de Mar Ibiza", 580, 400, "55", "clickandboat.com"),
    ("Ibiza", "Velero", "Oceanis 41.1", "12.4m", 8, "Port Ibiza", 700, 480, "65", "clickandboat.com"),
    ("Ibiza", "Velero", "Sun Odyssey 349", "10.3m", 6, "Marina Botafoc", 500, 350, "50", "clickandboat.com"),
    ("Menorca", "Velero", "Bavaria 37", "11.3m", 6, "Mahón", 480, 350, "50", "clickandboat.com"),
    ("Menorca", "Velero", "Dufour 382", "11.5m", 8, "Ciutadella", 520, 380, "55", "clickandboat.com"),
    ("Menorca", "Velero", "Oceanis 38.1", "11.5m", 8, "Mahón", 500, 360, "50", "clickandboat.com"),
    ("Menorca", "Velero", "Sun Odyssey 349", "10.3m", 6, "Mahón", 450, 320, "45", "clickandboat.com"),
]

headers_b = ["Isla", "Tipo", "Modelo", "Eslora", "Plazas", "Puerto Base",
             "Precio/día Ago (€)", "Precio/día Sep (€)", "Extras/día (€)", "Plataforma"]
ws2.append(headers_b)
style_header(ws2, 1, len(headers_b))

for i, row in enumerate(boat_data, start=2):
    ws2.append(list(row))
    f = LIGHT_GRAY if i % 2 == 0 else None
    style_row(ws2, i, len(headers_b), fill=f)

auto_width(ws2, len(headers_b))

# ════════════════════════════════════════════
# HOJA 3: COSTE TOTAL (Escenarios)
# ════════════════════════════════════════════
ws3 = wb.create_sheet("Coste Total")
ws3.sheet_properties.tabColor = "548235"

# Escenarios: combinación de mes, isla, días de barco
# Grupo: 3 Valencia, 1 Madrid, 1 Bolonia
scenarios = [
    # (Escenario, Mes, Isla, Días Barco, Coste VLC(3), Coste MAD(1), Coste BLQ(1), Coste Barco/día)
    # Agosto
    ("A1", "Agosto", "Ibiza", 2, 84, 112, 214, 500),
    ("A2", "Agosto", "Ibiza", 3, 84, 112, 214, 500),
    ("A3", "Agosto", "Menorca", 2, 131, 226, 286, 450),
    ("A4", "Agosto", "Menorca", 3, 131, 226, 286, 450),
    ("A5", "Agosto", "Ibiza", 2, 84, 130, 214, 500),
    # Septiembre
    ("S1", "Septiembre", "Ibiza", 2, 84, 67, 155, 350),
    ("S2", "Septiembre", "Ibiza", 3, 84, 67, 155, 350),
    ("S3", "Septiembre", "Menorca", 2, 105, 105, 195, 320),
    ("S4", "Septiembre", "Menorca", 3, 105, 105, 195, 320),
    ("S5", "Septiembre", "Ibiza", 2, 76, 70, 145, 350),
    ("S6", "Septiembre", "Ibiza", 3, 76, 70, 145, 350),
]

headers_c = ["Escenario", "Mes", "Isla", "Días Barco",
             "Vuelo VLC (x3) €", "Vuelo MAD (x1) €", "Vuelo BLQ (x1) €",
             "Coste Total Vuelos €", "Coste Barco Total €",
             "Coste Total Grupo €", "Coste por Persona €",
             "Ránking Economía"]
ws3.append(headers_c)
style_header(ws3, 1, len(headers_c))

# Sort scenarios by total cost per person
scenario_results = []
for s in scenarios:
    esc, mes, isla, dias, vlc, mad, blq, barco_dia = s
    total_vuelos = (vlc * 3) + mad + blq
    total_barco = barco_dia * dias
    total_grupo = total_vuelos + total_barco
    por_persona = round(total_grupo / 6)
    scenario_results.append((por_persona, esc, mes, isla, dias, vlc, mad, blq, total_vuelos, total_barco, total_grupo))

scenario_results.sort(key=lambda x: x[0])

for rank, (pp, esc, mes, isla, dias, vlc, mad, blq, tv, tb, tg) in enumerate(scenario_results, 1):
    r = rank + 1
    ws3.cell(row=r, column=1, value=f"{esc} (#{rank})")
    ws3.cell(row=r, column=2, value=mes)
    ws3.cell(row=r, column=3, value=isla)
    ws3.cell(row=r, column=4, value=dias)
    ws3.cell(row=r, column=5, value=vlc)
    ws3.cell(row=r, column=6, value=mad)
    ws3.cell(row=r, column=7, value=blq)
    ws3.cell(row=r, column=8, value=tv)
    ws3.cell(row=r, column=9, value=tb)
    ws3.cell(row=r, column=10, value=tg)
    ws3.cell(row=r, column=11, value=pp)
    ws3.cell(row=r, column=12, value=f"#{rank}")
    style_row(ws3, r, len(headers_c), fill=LIGHT_GRAY if r % 2 == 0 else None)
    # Highlight top 3
    if rank == 1:
        for c in range(1, len(headers_c) + 1):
            ws3.cell(row=r, column=c).fill = PatternFill(start_color="006100", end_color="006100", fill_type="solid")
            ws3.cell(row=r, column=c).font = Font(bold=True, color="FFFFFF")
    elif rank == 2:
        for c in range(1, len(headers_c) + 1):
            ws3.cell(row=r, column=c).fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
            ws3.cell(row=r, column=c).font = Font(color="FFFFFF")
    elif rank == 3:
        for c in range(1, len(headers_c) + 1):
            ws3.cell(row=r, column=c).fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
            ws3.cell(row=r, column=c).font = Font(color="FFFFFF")

auto_width(ws3, len(headers_c))

# Gráfico de barras
chart = BarChart()
chart.type = "col"
chart.title = "Coste por Persona por Escenario (€)"
chart.y_axis.title = "€ / persona"
chart.x_axis.title = "Escenario"
chart.style = 10

data_ref = Reference(ws3, min_col=11, min_row=1, max_row=len(scenario_results) + 1)
cats_ref = Reference(ws3, min_col=1, min_row=2, max_row=len(scenario_results) + 1)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.shape = 4
chart.width = 30
chart.height = 15

ws3.add_chart(chart, f"N2")

# ════════════════════════════════════════════
# HOJA 4: CALENDARIO
# ════════════════════════════════════════════
ws4 = wb.create_sheet("Calendario")
ws4.sheet_properties.tabColor = "BF8F00"

# Calendario visual para Agosto y Septiembre
months = [
    ("Agosto 2026", [
        (1, "Sáb"), (2, "Dom"), (3, "Lun"), (4, "Mar"), (5, "Mié"), (6, "Jue"), (7, "Vie"),
        (8, "Sáb"), (9, "Dom"), (10, "Lun"), (11, "Mar"), (12, "Mié"), (13, "Jue"), (14, "Vie"),
        (15, "Sáb"), (16, "Dom"), (17, "Lun"), (18, "Mar"), (19, "Mié"), (20, "Jue"), (21, "Vie"),
        (22, "Sáb"), (23, "Dom"), (24, "Lun"), (25, "Mar"), (26, "Mié"), (27, "Jue"), (28, "Vie"),
        (29, "Sáb"), (30, "Dom"), (31, "Lun")
    ]),
    ("Septiembre 2026", [
        (1, "Mar"), (2, "Mié"), (3, "Jue"), (4, "Vie"), (5, "Sáb"), (6, "Dom"), (7, "Lun"),
        (8, "Mar"), (9, "Mié"), (10, "Jue"), (11, "Vie"), (12, "Sáb"), (13, "Dom"), (14, "Lun"),
        (15, "Mar"), (16, "Mié"), (17, "Jue"), (18, "Vie"), (19, "Sáb"), (20, "Dom"), (21, "Lun"),
        (22, "Mar"), (23, "Mié"), (24, "Jue"), (25, "Vie"), (26, "Sáb"), (27, "Dom"), (28, "Lun"),
        (29, "Mar"), (30, "Mié")
    ])
]

row = 1
for month_name, days in months:
    ws4.cell(row=row, column=1, value=month_name).font = Font(bold=True, size=14, color="1F4E79")
    row += 1
    # Header: Día, Semana, Precio Vuelo, Disponibilidad
    headers_cal = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
    for c, h in enumerate(headers_cal, 1):
        ws4.cell(row=row, column=c, value=h)
        ws4.cell(row=row, column=c).font = Font(bold=True, size=10)
        ws4.cell(row=row, column=c).alignment = Alignment(horizontal="center")
        ws4.cell(row=row, column=c).fill = SUBHEADER_FILL
        ws4.cell(row=row, column=c).border = thin_border
    row += 1

    # Create a blank grid (weeks)
    first_day = days[0][1]
    day_names = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
    start_col = day_names.index(first_day) + 1

    cal_row = row
    for d, dn in days:
        c = day_names.index(dn) + 1
        cell = ws4.cell(row=cal_row, column=c, value=d)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        # Highlight weekend
        if dn in ("Sáb", "Dom"):
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        if dn == "Sáb":
            cal_row += 1

    # Add a note about best dates
    row = cal_row + 1
    ws4.cell(row=row, column=1, value="🏆 Mejores fechas (Septiembre, entre semana):").font = Font(bold=True, size=11, color="006100")
    row += 1
    ws4.cell(row=row, column=1, value="Martes 1 Sep - Jueves 3 Sep (3 días)").font = Font(size=10)
    row += 1
    ws4.cell(row=row, column=1, value="Martes 8 Sep - Jueves 10 Sep (3 días)").font = Font(size=10)
    row += 1
    ws4.cell(row=row, column=1, value="Martes 15 Sep - Jueves 17 Sep (3 días)").font = Font(size=10)
    row += 2

auto_width(ws4, 7, min_w=12, max_w=15)

# ════════════════════════════════════════════
# HOJA 5: RECOMENDACIÓN
# ════════════════════════════════════════════
ws5 = wb.create_sheet("Recomendación")
ws5.sheet_properties.tabColor = "006100"

rec_data = [
    ("RECOMENDACIÓN PRINCIPAL", ""),
    ("", ""),
    ("Destino:", "IBIZA (Ibiza)"),
    ("Mes:", "Septiembre 2026"),
    ("Fechas propuestas:", "Martes 1 - Jueves 3 de Septiembre (3 días)"),
    ("Días de barco:", "3 días"),
    ("", ""),
    ("DESGLOSE DE COSTES", ""),
    ("", ""),
    ("Vuelos Valencia → Ibiza (x3 personas):", "3 × €84 = €252"),
    ("  - Ryanair ida+vuelta desde €42/trayecto", ""),
    ("Vuelo Madrid → Ibiza (x1 persona):", "1 × €67 = €67"),
    ("  - Ryanair ida+vuelta desde €67", ""),
    ("Vuelo Bolonia → Ibiza (x1 persona):", "1 × €155 = €155"),
    ("  - Ryanair directo ida+vuelta", ""),
    ("Total vuelos:", "€474"),
    ("", ""),
    ("Barco (velero 6-8 pax, sin patrón):", "3 días × €350/día = €1.050"),
    ("  - Ej: Sun Odyssey 349 (6 pax) o Bavaria 40 (8 pax)", ""),
    ("  - Extras estimados (seguro+combustible): ~€50/día", ""),
    ("Total barco:", "€1.050 + €150 extras = €1.200"),
    ("", ""),
    ("", ""),
    ("COSTE TOTAL DEL VIAJE:", "€474 (vuelos) + €1.200 (barco) = €1.674"),
    ("COSTE POR PERSONA:", "€279"),
    ("", ""),
    ("DÓNDE RESERVAR:", ""),
    ("  Vuelos:", "Skyscanner / Ryanair (web oficial)"),
    ("  Barco:", "Click&Boat (clickandboat.com) - filtrar: velero, sin patrón, 6+ plazas"),
    ("", ""),
    ("NOTAS:", ""),
    ("  - Precios de vuelos confirmados vía Google Flights (consulta: May 2026)", ""),
    ("  - Precios de barcos estimados según mercado; consultar plataformas para precios exactos", ""),
    ("  - Septiembre es significativamente más barato que Agosto", ""),
    ("  - Entre semana (mar-jue) es más barato que fines de semana", ""),
]

for i, (label, value) in enumerate(rec_data, 1):
    ws5.cell(row=i, column=1, value=label).font = Font(bold=True, size=11) if ":" in label or label.isupper() else Font(size=11)
    ws5.cell(row=i, column=2, value=value).font = Font(size=11)
    if label == "RECOMENDACIÓN PRINCIPAL":
        ws5.cell(row=i, column=1).font = Font(bold=True, size=16, color="006100")
    elif label == "COSTE TOTAL DEL VIAJE:":
        ws5.cell(row=i, column=1).font = Font(bold=True, size=13, color="006100")
        ws5.cell(row=i, column=2).font = Font(bold=True, size=13, color="006100")
    elif label == "COSTE POR PERSONA:":
        ws5.cell(row=i, column=1).font = Font(bold=True, size=13, color="006100")
        ws5.cell(row=i, column=2).font = Font(bold=True, size=13, color="006100")
    elif label == "DESGLOSE DE COSTES":
        ws5.cell(row=i, column=1).font = Font(bold=True, size=13, color="1F4E79")

ws5.column_dimensions["A"].width = 45
ws5.column_dimensions["B"].width = 40

# ── Guardar ──
wb.save("C:\\dev\\viaje-barco\\informe-viaje-baleares.xlsx")
print("Excel generado: informe-viaje-baleares.xlsx")
